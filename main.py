#!/usr/bin/env python3
"""
HVTN300 V3 mAb Representative Selection Pipeline
=================================================
Orchestrates all five steps end-to-end from a single Excel input:

  Step 1  fasta.py      Excel → FASTA file of Fab sequences
  Step 2  ebi_tree.py   FASTA → EBI Clustal Omega alignment → EBI Simple
                         Phylogeny (UPGMA) → Newick tree + PNG
  Step 3  cluster_tree.py  Newick → UPGMA depth-cut clustering →
                            cluster_assignments.csv + coloured PNG
  Step 4  scalop.py     cluster_assignments.csv → SCALOP CDR extraction →
                         cluster_assignments_CDR.csv
  Step 5  PIM.py        cluster_assignments_CDR.csv → per-cluster Clustal Omega
                         alignment → PIM → ranked tables →
                         representatives_top_bottom.csv +
                         representatives_trisect.csv

All intermediate and final files are written to --out (default: ./results/).
The pipeline can be resumed from any step with --start-from.

Usage
-----
  # Full run (requires EBI internet access + SCALOP folder next to this script)
  python main.py --xlsx V3_mAb_sequences.xlsx --email you@example.com

  # Resume from clustering (tree already built)
  python main.py --xlsx V3_mAb_sequences.xlsx --email you@example.com \\
      --start-from cluster --newick results/tree.nwk

  # Resume from CDR extraction (clusters already assigned)
  python main.py --xlsx V3_mAb_sequences.xlsx --email you@example.com \\
      --start-from scalop --cluster-csv results/cluster_assignments.csv

  # Resume from PIM (CDR CSV already exists)
  python main.py --email you@example.com \\
      --start-from pim --cdr-csv results/cluster_assignments_CDR.csv

  # Force a specific number of clusters instead of auto-detection
  python main.py --xlsx V3_mAb_sequences.xlsx --email you@example.com --k 5

  # Choose 2 trisect representatives per cluster instead of 4
  python main.py --xlsx V3_mAb_sequences.xlsx --email you@example.com --n-reps 2
"""

import argparse
import logging
import sys
import os
from pathlib import Path

# ── logging ──────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)],
)
log = logging.getLogger(__name__)

STEPS = ["fasta", "tree", "cluster", "scalop", "pim"]

# SCALOP is expected to live next to this script (same folder)
# SCALOP is expected to be pip-installed into your conda environment:
#   conda activate hvtn300
#   pip install ./SCALOP/
# If for some reason it is NOT installed, pass --scalop-path to inject the path manually.
SCALOP_PATH = None   # None = use installed package (recommended)


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def _banner(msg: str) -> None:
    bar = "─" * 60
    log.info(bar)
    log.info("  %s", msg)
    log.info(bar)


def _require_file(path: str, label: str) -> None:
    if not Path(path).exists():
        log.error("Required file not found: %s  (%s)", path, label)
        sys.exit(1)


# ─────────────────────────────────────────────────────────────────────────────
# Step 1 — Excel → FASTA
# ─────────────────────────────────────────────────────────────────────────────

def step_fasta(xlsx: str, out_dir: str) -> str:
    """
    Convert Excel to FASTA. Supports both column layouts:
      • Simple:   Sequence Number | Heavy Chain | Light Chain
      • Full:     Ab ID | signed # | VH AA Sequence | VK/VL AA Sequence
    Returns path to the written FASTA file.
    """
    _banner("Step 1 — Excel → FASTA")
    import openpyxl
    import pandas as pd
    import re

    fasta_path = os.path.join(out_dir, "sequences.fasta")
    idmap_path = os.path.join(out_dir, "id_map.csv")

    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    log.info("Excel columns: %s", headers)

    # Detect column layout
    def _col(candidates):
        for name in candidates:
            if name in headers:
                return name
        return None

    vh_col  = _col(["Heavy Chain", "VH AA Sequence"])
    vl_col  = _col(["Light Chain", "VK/VL AA Sequence"])
    num_col = _col(["Sequence Number", "signed #"])

    if not vh_col or not vl_col:
        log.error(
            "Cannot find heavy/light chain columns. Found: %s\n"
            "Expected one of: 'Heavy Chain' / 'VH AA Sequence'  and  "
            "'Light Chain' / 'VK/VL AA Sequence'", headers
        )
        sys.exit(1)

    df = pd.read_excel(xlsx)
    df = df.dropna(subset=[vh_col, vl_col])

    records, id_rows = [], []
    for idx, (_, row) in enumerate(df.iterrows(), start=1):
        seq_id = f"Sequence_{idx}"
        vh = re.sub(r"[\s\n\r]+", "", str(row[vh_col]))
        vl = re.sub(r"[\s\n\r]+", "", str(row[vl_col]))
        if not vh or not vl:
            continue
        records.append((seq_id, vh + vl))
        id_rows.append({
            "Sequence_ID":       seq_id,
            "Sequence_Number":   idx,
            "Original_ID":       row.get(num_col, idx),
        })

    if not records:
        log.error("No valid Fab sequences found in %s", xlsx)
        sys.exit(1)

    with open(fasta_path, "w") as fh:
        for seq_id, seq in records:
            fh.write(f">{seq_id}\n")
            for i in range(0, len(seq), 60):
                fh.write(seq[i:i + 60] + "\n")

    pd.DataFrame(id_rows).to_csv(idmap_path, index=False)
    log.info("Wrote %d sequences → %s", len(records), fasta_path)
    log.info("ID map → %s", idmap_path)
    return fasta_path


# ─────────────────────────────────────────────────────────────────────────────
# Step 2 — FASTA → EBI tree → Newick + PNG
# ─────────────────────────────────────────────────────────────────────────────

def step_tree(xlsx: str, email: str, out_dir: str,
              clustering: str = "UPGMA") -> str:
    """
    Submit sequences to EBI Clustal Omega + Simple Phylogeny.
    Returns path to the saved Newick file.
    """
    _banner("Step 2 — EBI Clustal Omega + Simple Phylogeny → tree")
    import ebi_tree

    sequences = ebi_tree.load_sequences_from_xlsx(xlsx)
    if len(sequences) < 2:
        log.error("Need at least 2 sequences to build a tree.")
        sys.exit(1)

    alignment = ebi_tree.run_clustalo(sequences, email)
    newick    = ebi_tree.run_simple_phylogeny(alignment, email,
                                              clustering=clustering)

    aln_path    = os.path.join(out_dir, "alignment.aln")
    newick_path = os.path.join(out_dir, "tree.nwk")
    png_path    = os.path.join(out_dir, "tree.png")

    with open(aln_path, "w") as fh:
        fh.write(alignment)
    with open(newick_path, "w") as fh:
        fh.write(newick)

    ebi_tree.render_tree_png(newick, png_path)
    log.info("Alignment → %s", aln_path)
    log.info("Newick    → %s", newick_path)
    log.info("Tree PNG  → %s", png_path)
    return newick_path


# ─────────────────────────────────────────────────────────────────────────────
# Step 3 — Newick → cluster assignments CSV
# ─────────────────────────────────────────────────────────────────────────────

def step_cluster(newick_path: str, xlsx: str, out_dir: str,
                 k=None, k_min=2, k_max=20) -> str:
    """
    Cut the UPGMA tree into clusters and write cluster_assignments.csv.
    Returns path to the CSV.
    """
    _banner("Step 3 — UPGMA depth-cut clustering")
    import cluster_tree

    png_path = os.path.join(out_dir, "tree_clustered.png")
    cluster_tree.run(
        newick_path = newick_path,
        k           = k,
        k_min       = k_min,
        k_max       = k_max,
        out_dir     = out_dir,
        png_path    = png_path,
        xlsx_path   = xlsx,
    )
    csv_path = os.path.join(out_dir, "cluster_assignments.csv")
    _require_file(csv_path, "cluster_assignments.csv from cluster_tree")
    return csv_path


# ─────────────────────────────────────────────────────────────────────────────
# Step 4 — cluster assignments → CDR extraction
# ─────────────────────────────────────────────────────────────────────────────

def step_scalop(cluster_csv: str, out_dir: str, ncpu: int = 1,
                scalop_path: str = None) -> str:
    """
    Run SCALOP on every heavy/light pair and append CDR columns.
    Returns path to the CDR CSV.
    """
    _banner("Step 4 — SCALOP CDR extraction")
    import extract_cdrs

    cdr_csv = os.path.join(out_dir, "cluster_assignments_CDR.csv")
    extract_cdrs.run(
        csv_path    = cluster_csv,
        scalop_path = scalop_path,   # None = use pip-installed package
        out_path    = cdr_csv,
        ncpu        = ncpu,
    )
    _require_file(cdr_csv, "cluster_assignments_CDR.csv from extract_cdrs")
    return cdr_csv


# ─────────────────────────────────────────────────────────────────────────────
# Step 5 — CDR CSV → PIM → ranked tables + representatives
# ─────────────────────────────────────────────────────────────────────────────

def step_pim(cdr_csv: str, out_dir: str, n_reps: int = 4) -> None:
    """Run Clustal Omega PIM alignment and select representatives."""
    _banner("Step 5 — CDR PIM alignment + representative selection")
    import PIM

    PIM.run(
        csv_path = cdr_csv,
        out_dir  = out_dir,
        n_reps   = n_reps,
    )


# ─────────────────────────────────────────────────────────────────────────────
# Orchestrator
# ─────────────────────────────────────────────────────────────────────────────

def main():
    p = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )

    # ── Inputs ────────────────────────────────────────────────────────────────
    p.add_argument("--xlsx",  default=None,
                   help="Input Excel file (Sequence Number | Heavy Chain | Light Chain)")
    p.add_argument("--email", default=None,
                   help="Your email address — required by EBI REST API (steps 1–2)")

    # ── Resume handles ────────────────────────────────────────────────────────
    p.add_argument("--start-from",
                   choices=STEPS, default="fasta",
                   help="Resume from this step  [default: fasta]")
    p.add_argument("--newick",      default=None,
                   help="Existing .nwk file — required when --start-from cluster")
    p.add_argument("--cluster-csv", default=None,
                   help="Existing cluster_assignments.csv — required when "
                        "--start-from scalop")
    p.add_argument("--cdr-csv",     default=None,
                   help="Existing cluster_assignments_CDR.csv — required when "
                        "--start-from pim")

    # ── Clustering ────────────────────────────────────────────────────────────
    p.add_argument("--k",     type=int, default=None,
                   help="Force number of clusters (default: auto-detect)")
    p.add_argument("--k-min", type=int, default=2,
                   help="Min k for auto-detection  [default: 2]")
    p.add_argument("--k-max", type=int, default=20,
                   help="Max k for auto-detection  [default: 20]")

    # ── Tree method ───────────────────────────────────────────────────────────
    p.add_argument("--clustering-method", default="UPGMA",
                   choices=["UPGMA", "Neighbour-joining"],
                   help="Phylogenetic tree method  [default: UPGMA]")

    # ── Representatives ───────────────────────────────────────────────────────
    p.add_argument("--n-reps", type=int, default=4,
                   help="Trisect representatives per cluster  [default: 4]")

    # ── Parallelism ───────────────────────────────────────────────────────────
    p.add_argument("--ncpu", type=int, default=1,
                   help="Parallel workers for SCALOP  [default: 1]")
    p.add_argument("--scalop-path", default=None,
                   help="Path to SCALOP lib/python (only needed if scalop is NOT "
                        "pip-installed — normally leave this unset)")

    # ── Output ────────────────────────────────────────────────────────────────
    p.add_argument("--out", default="results",
                   help="Output directory  [default: results/]")

    args = p.parse_args()

    # ── Validate inputs ───────────────────────────────────────────────────────
    start_idx = STEPS.index(args.start_from)

    if start_idx <= STEPS.index("tree") and not args.email:
        p.error("--email is required for steps fasta and tree (EBI API calls)")

    if start_idx <= STEPS.index("tree") and not args.xlsx:
        p.error("--xlsx is required to run from steps fasta or tree")

    if args.start_from == "cluster":
        if not args.newick:
            p.error("--newick is required when --start-from cluster")
        _require_file(args.newick, "--newick")

    if args.start_from == "scalop":
        if not args.cluster_csv:
            p.error("--cluster-csv is required when --start-from scalop")
        _require_file(args.cluster_csv, "--cluster-csv")

    if args.start_from == "pim":
        if not args.cdr_csv:
            p.error("--cdr-csv is required when --start-from pim")
        _require_file(args.cdr_csv, "--cdr-csv")

    os.makedirs(args.out, exist_ok=True)
    log.info("Output directory: %s", os.path.abspath(args.out))
    log.info("Starting from step: %s", args.start_from)

    # ── Run steps ─────────────────────────────────────────────────────────────
    newick_path  = args.newick
    cluster_csv  = args.cluster_csv
    cdr_csv      = args.cdr_csv

    if start_idx <= STEPS.index("fasta"):
        # fasta.py is not directly called here — ebi_tree.py reads Excel natively,
        # so we just validate the Excel exists and let step_tree handle reading.
        _require_file(args.xlsx, "--xlsx")
        log.info("Step 1 (fasta): Excel validated — sequences will be read "
                 "directly by ebi_tree in Step 2")

    if start_idx <= STEPS.index("tree"):
        newick_path = step_tree(
            xlsx        = args.xlsx,
            email       = args.email,
            out_dir     = args.out,
            clustering  = args.clustering_method,
        )

    if start_idx <= STEPS.index("cluster"):
        cluster_csv = step_cluster(
            newick_path = newick_path,
            xlsx        = args.xlsx,
            out_dir     = args.out,
            k           = args.k,
            k_min       = args.k_min,
            k_max       = args.k_max,
        )

    if start_idx <= STEPS.index("scalop"):
        cdr_csv = step_scalop(
            cluster_csv = cluster_csv,
            out_dir     = args.out,
            ncpu        = args.ncpu,
            scalop_path = args.scalop_path,
        )

    if start_idx <= STEPS.index("pim"):
        step_pim(
            cdr_csv = cdr_csv,
            out_dir = args.out,
            n_reps  = args.n_reps,
        )

    # ── Summary ───────────────────────────────────────────────────────────────
    _banner("Pipeline complete")
    out = args.out
    files = {
        "Tree (Newick)":           os.path.join(out, "tree.nwk"),
        "Tree (PNG)":              os.path.join(out, "tree.png"),
        "Clustered tree (PNG)":    os.path.join(out, "tree_clustered.png"),
        "Cluster assignments":     os.path.join(out, "cluster_assignments.csv"),
        "CDR assignments":         os.path.join(out, "cluster_assignments_CDR.csv"),
        "PIM summary":             os.path.join(out, "PIM_summary.csv"),
        "Reps (top/bottom)":       os.path.join(out, "representatives_top_bottom.csv"),
        "Reps (trisect)":          os.path.join(out, "representatives_trisect.csv"),
    }
    log.info("Output files:")
    for label, path in files.items():
        status = "✓" if os.path.exists(path) else "–"
        log.info("  %s  %-30s  %s", status, label, path)


if __name__ == "__main__":
    main()

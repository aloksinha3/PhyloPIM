#!/usr/bin/env python3
"""
CDR Loop Extraction for Clustered Antibody Sequences
=====================================================
Reads cluster_assignments.csv (output of cluster_tree.py), runs SCALOP on
each heavy/light chain pair to extract all 6 CDR loops, and writes a new
CSV with CDR sequence and residue position columns appended.

Improvements over run_scalop_cdr.py
-------------------------------------
- Works directly from cluster_assignments.csv — no Excel dependency or
  hardcoded column numbers.
- Processes sequences in parallel (--ncpu) for speed.
- Optional --ground-truth flag to validate SCALOP output against a
  pre-computed Excel file (e.g. V3_mAb_NSEM_CDR.xlsx) and report mismatches.
- Separates sequence and position into distinct columns (CDRH1_seq,
  CDRH1_pos) so downstream analysis can filter/sort on either independently.
- Graceful per-sequence error handling: failures are logged without
  stopping the run.

Usage
-----
  python extract_cdrs.py --csv cluster_assignments.csv --scalop /path/to/SCALOP/lib/python
  python extract_cdrs.py --csv cluster_assignments.csv --scalop /path/to/SCALOP/lib/python \\
      --out cluster_assignments_CDR.csv
  python extract_cdrs.py --csv cluster_assignments.csv --scalop /path/to/SCALOP/lib/python \\
      --ground-truth V3_mAb_NSEM_CDR.xlsx
  python extract_cdrs.py --csv cluster_assignments.csv --scalop /path/to/SCALOP/lib/python \\
      --ncpu 4
"""

import argparse
import copy
import logging
import re
import sys
from concurrent.futures import ProcessPoolExecutor, as_completed
from pathlib import Path

import pandas as pd

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)],
)
log = logging.getLogger(__name__)

CDR_NAMES = ["CDRH1", "CDRH2", "CDRH3", "CDRL1", "CDRL2", "CDRL3"]
CDR_KEYS  = ["H1",    "H2",    "H3",    "L1",    "L2",    "L3"]
SCHEME     = "imgt"
DEFINITION = "north"


# ─────────────────────────────────────────────────────────────────────────────
# SCALOP helpers
# ─────────────────────────────────────────────────────────────────────────────

def _import_scalop(scalop_path: str = None):
    """
    Import SCALOP. Two modes:
      1. scalop_path=None  — use the already pip-installed scalop package
                             (works after: pip install ./SCALOP/)
      2. scalop_path=str   — inject path into sys.path first, then import
                             (only needed if scalop was NOT pip-installed)
    """
    if scalop_path and scalop_path not in sys.path:
        sys.path.insert(0, scalop_path)
    try:
        from scalop.anarci import run_anarci
        from scalop.utils import getnumberedCDRloop
        return run_anarci, getnumberedCDRloop
    except ImportError as e:
        hint = f"from '{scalop_path}'" if scalop_path else "from the installed environment"
        raise ImportError(
            f"Could not import SCALOP {hint}.\n"
            f"Fix: activate your conda env and run:  pip install ./SCALOP/\n"
            f"Original error: {e}"
        )


def _clean(seq) -> str:
    if seq is None:
        return ""
    return re.sub(r"[\s\n\r]+", "", str(seq).strip())


def extract_cdrs(seq_id: str, heavy: str, light: str,
                 scalop_path: str) -> dict:
    """
    Run SCALOP on one heavy/light pair.
    Returns a flat dict:
      { "CDRH1_seq": "KASGDTFGSN", "CDRH1_pos": "23-32", ... }
    Failures are recorded as empty strings with a warning.
    """
    run_anarci, getnumberedCDRloop = _import_scalop(scalop_path)

    heavy = _clean(heavy)
    light = _clean(light)
    result = {}

    if not heavy or not light:
        log.warning("%s: missing heavy or light chain — skipping", seq_id)
        for name in CDR_NAMES:
            result[f"{name}_seq"] = ""
            result[f"{name}_pos"] = ""
        return result

    seqs = [("heavy", heavy), ("light", light)]
    try:
        numbered = run_anarci(seqs, scheme=SCHEME, ncpu=1, assign_germline=False)
    except Exception as e:
        log.warning("%s: ANARCI failed — %s", seq_id, e)
        for name in CDR_NAMES:
            result[f"{name}_seq"] = ""
            result[f"{name}_pos"] = f"Error: {e}"
        return result

    if numbered[2][0] is None or numbered[2][1] is None:
        log.warning("%s: ANARCI returned None for one chain", seq_id)
        for name in CDR_NAMES:
            result[f"{name}_seq"] = ""
            result[f"{name}_pos"] = "ANARCI failed"
        return result

    chain_idx = {"H1": 0, "H2": 0, "H3": 0, "L1": 1, "L2": 1, "L3": 1}

    for cdr_name, cdr_key in zip(CDR_NAMES, CDR_KEYS):
        idx      = chain_idx[cdr_key]
        num      = copy.deepcopy(numbered[1][idx][0][0])
        loop, _  = getnumberedCDRloop(num, cdr_key, SCHEME, DEFINITION)

        if loop:
            # 1-based residue positions in the original (non-gapped) sequence
            seq_ab  = sorted([x for x in num if x[1] != "-"])
            start1  = seq_ab.index(loop[0])  + 1
            end1    = seq_ab.index(loop[-1]) + 1
            cdr_seq = "".join(x[1] for x in loop)
            result[f"{cdr_name}_seq"] = cdr_seq
            result[f"{cdr_name}_pos"] = f"{start1}-{end1}"
        else:
            result[f"{cdr_name}_seq"] = ""
            result[f"{cdr_name}_pos"] = ""

    return result


def _worker(args):
    """Top-level function required for multiprocessing pickling."""
    seq_id, heavy, light, scalop_path = args
    return seq_id, extract_cdrs(seq_id, heavy, light, scalop_path)


# ─────────────────────────────────────────────────────────────────────────────
# Optional: validate against ground truth Excel
# ─────────────────────────────────────────────────────────────────────────────

def load_ground_truth(xlsx_path: str) -> pd.DataFrame:
    """
    Load CDR ground truth from V3_mAb_NSEM_CDR.xlsx.
    Col 2 = sequence number (integer), cols 36-41 = CDR strings "SEQ (start-end)".
    Returns DataFrame with Seq_ID and one column per CDR (sequence only).
    """
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active

    rows = []
    for r in range(2, ws.max_row + 1):
        seq_num = ws.cell(r, 2).value
        if seq_num is None:
            continue
        row = {"Seq_ID": f"Sequence_{int(seq_num)}"}
        for i, name in enumerate(CDR_NAMES):
            val = ws.cell(r, 36 + i).value or ""
            # Ground truth format: "SEQUENCE (start-end)" — extract sequence only
            match = re.match(r"^([A-Z]+)\s*\(", str(val))
            row[f"{name}_gt"] = match.group(1) if match else str(val).strip()
        rows.append(row)

    return pd.DataFrame(rows).set_index("Seq_ID")


def compare_to_ground_truth(df: pd.DataFrame, gt: pd.DataFrame) -> None:
    """Log per-CDR match rates and list any mismatches."""
    shared = df.index.intersection(gt.index)
    if shared.empty:
        log.warning("No shared Seq_IDs between results and ground truth.")
        return

    log.info("Comparing %d sequences against ground truth …", len(shared))
    total_mismatches = 0

    for name in CDR_NAMES:
        pred_col = f"{name}_seq"
        gt_col   = f"{name}_gt"
        if pred_col not in df.columns or gt_col not in gt.columns:
            continue

        matches = (df.loc[shared, pred_col] == gt.loc[shared, gt_col]).sum()
        n       = len(shared)
        log.info("  %s: %d/%d match (%.0f%%)", name, matches, n, 100 * matches / n)

        mismatches = shared[df.loc[shared, pred_col] != gt.loc[shared, gt_col]]
        for sid in mismatches:
            pred = df.loc[sid, pred_col]
            gt_v = gt.loc[sid, gt_col]
            log.debug("    MISMATCH %s %s: SCALOP=%r  GT=%r", sid, name, pred, gt_v)
            total_mismatches += 1

    if total_mismatches == 0:
        log.info("All CDR sequences match ground truth exactly.")
    else:
        log.info("%d total mismatches (run with --debug to see details)", total_mismatches)


# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────

def run(csv_path: str, scalop_path: str, out_path: str = None,
        ncpu: int = 1, ground_truth: str = None) -> pd.DataFrame:

    df = pd.read_csv(csv_path)
    required = {"Seq_ID", "Heavy_chain", "Light_chain"}
    missing  = required - set(df.columns)
    if missing:
        raise ValueError(
            f"cluster_assignments.csv is missing columns: {missing}\n"
            f"Found: {list(df.columns)}\n"
            f"Make sure you are using the output of cluster_tree.py."
        )

    log.info("Loaded %d sequences from %s", len(df), csv_path)

    # ── Run SCALOP ────────────────────────────────────────────────────────────
    tasks = [
        (row.Seq_ID, row.Heavy_chain, row.Light_chain, scalop_path)
        for row in df.itertuples()
    ]

    cdr_results = {}
    if ncpu > 1:
        log.info("Running SCALOP on %d sequences with %d workers …", len(tasks), ncpu)
        with ProcessPoolExecutor(max_workers=ncpu) as ex:
            futures = {ex.submit(_worker, t): t[0] for t in tasks}
            done = 0
            for future in as_completed(futures):
                seq_id, result = future.result()
                cdr_results[seq_id] = result
                done += 1
                if done % 10 == 0:
                    log.info("  %d / %d done …", done, len(tasks))
    else:
        log.info("Running SCALOP on %d sequences (single process) …", len(tasks))
        for i, task in enumerate(tasks, 1):
            seq_id, result = _worker(task)
            cdr_results[seq_id] = result
            if i % 10 == 0:
                log.info("  %d / %d done …", i, len(tasks))

    # ── Append CDR columns to DataFrame ──────────────────────────────────────
    cdr_df = pd.DataFrame.from_dict(cdr_results, orient="index")
    cdr_df.index.name = "Seq_ID"

    # Column order: seq then pos for each CDR
    col_order = [f"{n}_{s}" for n in CDR_NAMES for s in ("seq", "pos")]
    cdr_df = cdr_df[[c for c in col_order if c in cdr_df.columns]]

    result_df = df.set_index("Seq_ID").join(cdr_df).reset_index()

    # ── Validate against ground truth ─────────────────────────────────────────
    if ground_truth:
        log.info("Loading ground truth from %s …", ground_truth)
        gt = load_ground_truth(ground_truth)
        compare_to_ground_truth(result_df.set_index("Seq_ID"), gt)

    # ── Build concatenated CDR column (all 6 CDR sequences) ──────────────────
    seq_cols = [f"{name}_seq" for name in CDR_NAMES if f"{name}_seq" in result_df.columns]
    if seq_cols:
        def _concat_cdrs(row):
            parts = []
            for col in seq_cols:
                val = row.get(col, "")
                if pd.isna(val):
                    continue
                s = str(val).strip().replace(" ", "")
                if s:
                    parts.append(s)
            return "".join(parts)

        result_df["CDR_concat"] = result_df.apply(_concat_cdrs, axis=1)

    # ── Save ──────────────────────────────────────────────────────────────────
    if out_path is None:
        out_path = str(Path(csv_path).with_suffix("")) + "_CDR.csv"

    result_df.to_csv(out_path, index=False)
    log.info("Saved → %s  (%d rows × %d columns)", out_path, len(result_df), len(result_df.columns))
    return result_df


# ─────────────────────────────────────────────────────────────────────────────
# CLI
# ─────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    p = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    p.add_argument("--csv",          required=True,
                   help="cluster_assignments.csv from cluster_tree.py")
    p.add_argument("--scalop",       required=True,
                   help="Path to SCALOP lib/python directory "
                        "(e.g. /Users/you/Desktop/SCALOP/lib/python)")
    p.add_argument("--out",          default=None,
                   help="Output CSV path  [default: <csv>_CDR.csv]")
    p.add_argument("--ncpu",         type=int, default=1,
                   help="Parallel workers  [default: 1]")
    p.add_argument("--ground-truth", default=None,
                   help="V3_mAb_NSEM_CDR.xlsx to validate SCALOP output against")
    p.add_argument("--debug",        action="store_true",
                   help="Show per-mismatch details in ground-truth comparison")
    args = p.parse_args()

    if args.debug:
        logging.getLogger().setLevel(logging.DEBUG)

    run(
        csv_path     = args.csv,
        scalop_path  = args.scalop,
        out_path     = args.out,
        ncpu         = args.ncpu,
        ground_truth = args.ground_truth,
    )
